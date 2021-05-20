import sqlite3
import eel
import os
from urllib.request import urlopen
from bs4 import BeautifulSoup
from datetime import *
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side

class Database:
    def __init__(self, db):
        self.conn = sqlite3.connect(db)
        self.cur = self.conn.cursor()
        self.cur.execute(
            """CREATE TABLE IF NOT EXISTS staff 
            (id INTEGER PRIMARY KEY, 
            nick text, 
            rank text, 
            vacation text,
            time mediumtext)
            """)
        self.conn.commit()
    def create(self, nick, rank, vacation, time):
        self.cur.execute("INSERT INTO staff VALUES (NULL, ?, ?, ?, ?)",
                         (nick, rank, vacation, time))
        self.conn.commit()

    def read(self):
        self.cur.execute("SELECT * FROM staff")
        rows = self.cur.fetchall()
        return rows

    def update(self, id, nick, rank, vacation, time):
        self.cur.execute("UPDATE staff SET nick = ?, rank = ?, vacation = ?, time = ? WHERE id = ?",
                         (nick, rank, vacation, time, id))
        self.conn.commit()

    def change_id(self, nick, id):
        self.cur.execute("UPDATE staff SET id = ? WHERE nick =?",
                         (id, nick))
        self.conn.commit()

    def delete(self, id):
        self.cur.execute("DELETE FROM staff WHERE id=?", (id,))
        self.conn.commit()

    def delete_all(self):
        self.cur.execute("DELETE FROM staff")
        self.conn.commit()

    def __del__(self):
        self.conn.close()

class Database2:
    def __init__(self, db):
        self.conn = sqlite3.connect(db)
        self.cur = self.conn.cursor()
        self.cur.execute(
            """CREATE TABLE IF NOT EXISTS app_data 
            (id INTEGER PRIMARY KEY, 
            admin_id text, 
            time_result text)
            """)
        self.conn.commit()
    def create(self, admin_id, time_result):
        self.cur.execute("INSERT INTO app_data VALUES (NULL, ?, ?)",
                         (admin_id, time_result))
        self.conn.commit()

    def read(self):
        self.cur.execute("SELECT * FROM app_data")
        rows = self.cur.fetchall()
        return rows

    def update(self, id, admin_id, time_result):
        self.cur.execute("UPDATE app_data SET admin_id = ?, time_result = ? WHERE id = ?",
                         (admin_id, time_result, id))
        self.conn.commit()

    def __del__(self):
        self.conn.close()

db = Database('staff.db')
db_app = Database2('app_data.db')

@eel.expose
def check_db():
    staff_list = db.read()
    if len(staff_list) == 0:
        eel.no_db('База данных с персоналом ещё не создана.')
    else:
        for i in staff_list:
            for j in i:
                if i.index(j) in [0, 1, 2]:
                    eel.db(str(j) + ' ')
                elif i.index(j) == 4:
                    if j.find(',') != -1:
                        j = j.split(',')
                        j = j[-1]
                        j = j[j.find('-') + 1:]
                        eel.db(j)
                    else:
                        j = j[j.find('-') + 1:]
                        eel.db(j)
            eel.db('\n')

@eel.expose
def fill_db():
    if len(db.read()) == 0:
        page = urlopen('http://deskt.fastrp.ru/staff/')
        soup = BeautifulSoup(page, 'html.parser')
        tr_tags = soup.find_all('tr')
        array = []
        today = date.today()
        today_date = f'{today.day}.{today.month}.{today.year}'
        for i in tr_tags:
            for j in i:
                if j.name != 'th' and i.index(j) == 1:
                    steam_id = j.a.string
                    j = str(j)
                    nick = j[len('<td data-label="Никнейм">'):j.find('(')]
                    array.append(f'{nick}({steam_id})')
                elif j.name != 'th' and i.index(j) == 2:
                    array.append(j.string)
                elif j.name != 'th' and i.index(j) == 3:
                    array.append(f'{today_date}-{j.string}')
            if len(array) == 3:
                db.create(array[0], array[1], '', array[2])
            array = []
        check_db()

@eel.expose
def update_db():
    if len(db.read()) > 0:
        page = urlopen('http://deskt.fastrp.ru/staff/')
        soup = BeautifulSoup(page, 'html.parser')
        tr_tags = soup.find_all('tr')
        staff_list = db.read()
        array = []
        steam_id_list = []
        steam_id = ''
        today = date.today()
        today_date = f'{today.day}.{today.month}.{today.year}'
        for i in staff_list:
            steam_id_list_item = i[1]
            steam_id_list_item = steam_id_list_item[steam_id_list_item.find('(') + 1:steam_id_list_item.find(')')]
            steam_id_list.append(steam_id_list_item)
        for i in tr_tags:
            for j in i:
                if j.name != 'th' and i.index(j) == 1:
                    steam_id = j.a.string
                    j = str(j)
                    nick = j[len('<td data-label="Никнейм">'):j.find('(')]
                    array.append(f'{nick}({steam_id})')
                elif j.name != 'th' and i.index(j) == 2:
                    array.append(j.string)
                elif j.name != 'th' and i.index(j) == 3:
                    array.append(f'{today_date}-{j.string}')
            if len(array) == 3:
                if steam_id not in steam_id_list:
                    db.create(array[0], array[1], '', array[2])
                    break
            array = []
        check_db()

@eel.expose
def get_time():
    if len(db.read()) > 0:
        page = urlopen('http://deskt.fastrp.ru/staff/')
        soup = BeautifulSoup(page, 'html.parser')
        tr_tags = soup.find_all('tr')
        staff_list = db.read()
        array = []
        today = date.today()
        today_date = f'{today.day}.{today.month}.{today.year}'
        for i in tr_tags:
            for j in i:
                if j.name != 'th' and i.index(j) == 1:
                    array.append(j.a.string)
                elif j.name != 'th' and i.index(j) == 3:
                    array.append(f'{today_date}-{j.string}')
            if len(array) == 2:
                for i in staff_list:
                    name = i[1]
                    steam_id = name[name.find('(') + 1:name.find(')')]
                    if steam_id == array[0]:
                        db.update(i[0], i[1], i[2], i[3], f'{i[4]},{array[1]}')
                        break
            array = []
        staff_list = db.read()
        array = []
        for i in staff_list:
            time = i[4]
            time = time.split(',')
            for j in time:
                time_date = j[:j.find('-')]
                if time_date not in [k[:k.find('-')] for k in array]:
                    array.append(j)
            time = ','.join(array)
            db.update(i[0], i[1], i[2], i[3], time)
            array = []
        check_db()

@eel.expose
def delete_db():
    if len(db.read()) > 0:
        db.delete_all()
        check_db()

@eel.expose
def change_admin(admin_id, admin_new_nick, admin_new_rank):
    if len(db.read()) > 0:
        staff_list = db.read()
        admin_id = int(admin_id)
        index = admin_id - 1
        admin_name = staff_list[index][1]
        if len(admin_new_nick) == 0:
            admin_new_nick = admin_name[:admin_name.find('(')]
        if len(admin_new_rank) == 0:
            admin_new_rank = staff_list[index][2]
        steam_id = admin_name[admin_name.find('('):]
        admin_new_name = admin_new_nick + steam_id
        db.update(staff_list[index][0], admin_new_name, admin_new_rank, staff_list[index][3], staff_list[index][4])
        check_db()

@eel.expose
def delete_admin(admin_id):
    db.delete(admin_id)
    staff_list = db.read()
    max_id = len(staff_list)
    for i in range(max_id):
        if staff_list[i][0] != i + 1:
            db.change_id(staff_list[i][1], i + 1)
    check_db()

@eel.expose
def check_admin_time():
    app_data = db_app.read()
    if len(db.read()) > 0:
        if len(app_data) > 0 and app_data[0][1] != '':
            eel.no_admin_time(f'Выбранный администратор: {db.read()[int(app_data[0][1]) - 1][1]}')
        else:
            eel.no_admin_time('Администратор ещё не был выбран.')

@eel.expose
def check_time(admin_id):
    if len(db.read()) > 0:
        app_data = db_app.read()
        if len(app_data) == 0:
            db_app.create(admin_id, '')
        elif app_data[0][1] != admin_id:
            db_app.update(app_data[0][0], admin_id, '')
        else:
            db_app.update(app_data[0][0], admin_id, app_data[0][2])
        admin_id = int(admin_id)
        admin_time = db.read()[admin_id - 1][4]
        admin_time = admin_time.split(',')
        admin_time_date = [i[:i.find('-')] for i in admin_time]
        for i in admin_time_date:
            eel.admin_time(i + '\n')

@eel.expose
def check_time_2():
    app_data = db_app.read()
    if len(app_data) > 0 and app_data[0][1] != '':
        admin_id = app_data[0][1]
        check_time(admin_id)

@eel.expose
def admin_week_time(date_1, date_2):
    if len(db.read()) > 0 and len(db_app.read()) > 0:
        app_data = db_app.read()
        admin_id = app_data[0][1]
        admin_id = int(admin_id)
        admin_time = db.read()[admin_id - 1][4]
        admin_time = admin_time.split(',')
        admin_time_date = [i[:i.find('-')] for i in admin_time]
        admin_time = [i[i.find('-') + 1:] for i in admin_time]
        admin_time = [int(i[:i.find(':')]) for i in admin_time]
        time_1 = admin_time[admin_time_date.index(date_1)]
        time_2 = admin_time[admin_time_date.index(date_2)]
        admin_time_res = time_2 - time_1
        admin_time_res = str(admin_time_res)
        if int(admin_time_res) < 0:
            admin_time_res = 'ошибка. Дата слева не должна быть больше даты справа.'
        db_app.update(app_data[0][0], app_data[0][1], f'{admin_time_res}:{date_1}-{date_2}')
        eel.admin_time_res(admin_time_res, date_1, date_2)

@eel.expose
def admin_week_time_2():
    app_data = db_app.read()
    if len(app_data) > 0 and app_data[0][2] != '':
        admin_time_res = app_data[0][2][:app_data[0][2].find(':')]
        two_dates = app_data[0][2][app_data[0][2].find(':') + 1:]
        date_1 = two_dates[:two_dates.find('-')]
        date_2 = two_dates[two_dates.find('-') + 1:]
        eel.admin_time_res(admin_time_res, date_1, date_2)

@eel.expose
def check_vacation():
    admin_id = int(db_app.read()[0][1]) - 1
    if db.read()[admin_id][3] != '':
        vacation = db.read()[admin_id][3]
        start_vacation = vacation[:vacation.find('-')]
        end_vacation = vacation[vacation.find('-') + 1:]
        eel.get_vacation(start_vacation, end_vacation)

@eel.expose
def give_vacation(start_vacation, end_vacation):
    if len(db.read()) > 0:
        admin_id = int(db_app.read()[0][1]) - 1
        staff_list = db.read()
        start_vacation = datetime.strptime(start_vacation, '%Y-%m-%d')
        start_vacation = f'{start_vacation.day}.{start_vacation.month}.{start_vacation.year}'
        end_vacation = datetime.strptime(end_vacation, '%Y-%m-%d')
        end_vacation = f'{end_vacation.day}.{end_vacation.month}.{end_vacation.year}'
        vacation = f'{start_vacation}-{end_vacation}'
        db.update(staff_list[admin_id][0], staff_list[admin_id][1], staff_list[admin_id][2], vacation,
                  staff_list[admin_id][4])

@eel.expose
def delete_vacation():
    if len(db.read()) > 0:
        staff_list = db.read()
        admin_id = int(db_app.read()[0][1]) - 1
        db.update(staff_list[admin_id][0], staff_list[admin_id][1], staff_list[admin_id][2], '',
                  staff_list[admin_id][4])

@eel.expose
def date_list():
    staff_list = db.read()
    date_list = ''
    max_value = 0
    for i in staff_list:
        if len(i[4]) > max_value:
            max_value = len(i[4])
            date_list = i[4]
    date_list = date_list.split(',')
    date_list = [i[:i.find('-')] for i in date_list]
    return date_list

@eel.expose
def create_excel_prep():
    today = date.today()
    week_ago = date.today() - timedelta(days=7)
    date_1 = f'{week_ago.day}.{week_ago.month}.{week_ago.year}'
    date_2 = f'{today.day}.{today.month}.{today.year}'
    create_excel(date_1, date_2)

@eel.expose
def create_excel(date_1, date_2):
    wb = Workbook()
    ws = wb.create_sheet(title='staff')
    def set_border(ws, cell_range):
        rows = ws[cell_range]
        for row in rows:
            if row == rows[0][0] or row == rows[0][-1] or row == rows[-1][0] or row == rows[-1][-1]:
                pass
            else:
                row[0].border = Border(left=Side(style='thin'))
                row[-1].border = Border(right=Side(style='thin'))
            for c in rows[0]:
                c.border = Border(top=Side(style='thin'))
            for c in rows[-1]:
                c.border = Border(bottom=Side(style='thin'))
        rows[0][0].border = Border(left=Side(style='thin'), top=Side(style='thin'))
        rows[0][-1].border = Border(right=Side(style='thin'), top=Side(style='thin'))
        rows[-1][0].border = Border(left=Side(style='thin'), bottom=Side(style='thin'))
        rows[-1][-1].border = Border(right=Side(style='thin'), bottom=Side(style='thin'))

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    def create_column(ws, cells, string, color):
        cell = cells[:cells.find(':')]
        if cells.find(':') == -1:
            ws[f'{cells}'] = f'{string}'
            ws[f'{cells}'].alignment = Alignment(horizontal='center')
            ws[f'{cells}'].fill = PatternFill(start_color=f'{color}', fill_type='solid')
            ws[f'{cells}'].border = thin_border
        else:
            ws.merge_cells(f'{cells}')
            ws[f'{cell}'] = f'{string}'
            ws[f'{cell}'].alignment = Alignment(horizontal='center')
            ws[f'{cell}'].fill = PatternFill(start_color=f'{color}', fill_type='solid')
            set_border(ws, f'{cells}')

    staff_list = db.read()
    max_id = len(staff_list)
    end = max_id + 6
    for i in range(1, end + 9):
        if i != end + 6:
            create_column(ws, f'A{i}:D{i}', '', 'ffffff')
    for i in range(1, end + 6):
        create_column(ws, f'E{i}:F{i}', '', 'ffffff')
    for i in range(1, end):
        create_column(ws, f'G{i}:H{i}', '', 'ffffff')
    for i in range(1, end):
        create_column(ws, f'I{i}:J{i}', '', 'ffffff')
    for i in range(1, end):
        create_column(ws, f'K{i}:M{i}', '', 'ffffff')
    for i in range(1, end):
        create_column(ws, f'N{i}:S{i}', '', 'ffffff')
    create_column(ws, f'A{end + 6}:S{end + 6}', '', 'ffffff')

    create_column(ws, 'A1:D1', 'Персонал', 'd5e6c6')
    create_column(ws, 'E1:F1', date_1, 'd5e6c6')
    create_column(ws, 'G1:H1', date_2, 'd5e6c6')
    create_column(ws, 'I1:J1', 'Часы', 'd5e6c6')
    create_column(ws, 'K1:M1', 'Отпуск', 'd5e6c6')
    create_column(ws, 'N1:S1', 'Заметки', 'd5e6c6')
    def create_end_row(order, bg_color, num, meaning):
        if num != '':
            create_column(ws, f'A{end + order}:D{end + order}', '', bg_color)
            create_column(ws, f'E{end + order}:F{end + order}', num, 'ffffff')
            create_column(ws, f'G{end + order}:S{end + order}', meaning, 'ffffff')
        else:
            create_column(ws, f'A{end + order}:D{end + order}', '', bg_color)
            create_column(ws, f'E{end + order}:S{end + order}', meaning, 'ffffff')
    create_end_row(0, 'e74c3c', 'Велика', 'Cупер администратор')
    create_end_row(1, '228b22', 'Свиняка', 'Оператор')
    create_end_row(2, '3a46f6', '4', 'Старший администратор')
    create_end_row(3, 'ff9900', '3', 'Администратор')
    create_end_row(4, '00e7ff', '2', 'Модератор')
    create_end_row(5, 'ff00d4', '1', 'Хелпер')
    create_end_row(7, '00b0f0', '', 'Отпуск')
    create_end_row(8, 'ffff00', '', 'Очень мало часиков')

    sa_list = []
    operator_list = []
    st_list = []
    admin_list = []
    moderator_list = []
    helper_list = []
    for i in staff_list:
        if i[2] == 'Куратор':
            operator_list.append(i)
        elif i[2] == 'Cупер администратор':
            sa_list.append(i)
        elif i[2] == 'Старший администратор':
            st_list.append(i)
        elif i[2] == 'Администратор':
            admin_list.append(i)
        elif i[2] == 'Модератор':
            moderator_list.append(i)
        elif i[2] == 'Помощник':
            helper_list.append(i)

    def create_row(date_1, date_2, list_name, index, color):
        date_1_date_type = datetime.strptime(date_1, '%d.%m.%Y')
        date_2_date_type = datetime.strptime(date_2, '%d.%m.%Y')
        time_1 = '?'
        time_2 = '?'
        create_column(ws, f'A{index}:D{index}', f'{list_name[0][1]}', color)
        time_all = list_name[0][4]
        time_all = time_all.split(',')
        for i in time_all:
            if i[:i.find('-')] == date_1:
                time_1 = i[i.find('-') + 1:]
                time_1 = time_1[:time_1.find(':')]
            elif i[:i.find('-')] == date_2:
                time_2 = i[i.find('-') + 1:]
                time_2 = time_2[:time_2.find(':')]
        if time_1 != '?' and time_1 != '?':
            time_1 = int(time_1)
            time_2 = int(time_2)
            time_delta = time_2 - time_1
        else:
            time_delta = '?'
        if len(list_name[0][3]) > 0:
            vacation = list_name[0][3]
            vacation_last_day = vacation[vacation.find('-') + 1:]
            vacation_first_day = vacation[:vacation.find('-')]
            vacation_last_day_date = datetime.strptime(vacation_last_day, '%d.%m.%Y')
            if date_2_date_type - vacation_last_day_date < date_2_date_type - date_1_date_type:
                create_column(ws, f'E{index}:F{index}', time_1, '00b0f0')
                create_column(ws, f'G{index}:H{index}', time_2, '00b0f0')
                create_column(ws, f'I{index}:J{index}', time_delta, '00b0f0')
                create_column(ws, f'K{index}:M{index}', f'+({vacation_first_day}-{vacation_last_day})', '00b0f0')
            else:
                create_column(ws, f'E{index}:F{index}', time_1, 'ffffff')
                create_column(ws, f'G{index}:H{index}', time_2, 'ffffff')
                create_column(ws, f'I{index}', time_delta, 'ffffff')
        elif type(time_delta) == int and time_delta < 5:
            create_column(ws, f'E{index}:F{index}', time_1, 'ffff00')
            create_column(ws, f'G{index}:H{index}', time_2, 'ffff00')
            create_column(ws, f'I{index}:J{index}', time_delta, 'ffff00')
        else:
            create_column(ws, f'E{index}:F{index}', time_1, 'ffffff')
            create_column(ws, f'G{index}:H{index}', time_2, 'ffffff')
            create_column(ws, f'I{index}:J{index}', time_delta, 'ffffff')
        list_name.pop(0)

    sa_admins = len(sa_list)
    for i in range(2, max_id + 2):
        if len(sa_list) > 0:
            create_row(date_1, date_2, sa_list, i, 'e74c3c')
            continue
        elif len(operator_list) > 0:
            create_row(date_1, date_2, operator_list, i, '228b22')
            continue
        elif len(st_list) > 0:
            create_row(date_1, date_2, st_list, i, '3a46f6')
            continue
        elif len(admin_list) > 0:
            create_row(date_1, date_2, admin_list, i, 'ff9900')
            continue
        elif len(moderator_list) > 0:
            create_row(date_1, date_2, moderator_list, i, '00e7ff')
            continue
        elif len(helper_list) > 0:
            create_row(date_1, date_2, helper_list, i, 'ff00d4')
            continue
    for i in range(2, sa_admins + 2):
        create_column(ws, f'E{i}:F{i}', '-', 'ffffff')
        create_column(ws, f'G{i}:H{i}', '-', 'ffffff')
        create_column(ws, f'I{i}:J{i}', '-', 'ffffff')

    today = date.today()
    today_date = f'{today.day}.{today.month}.{today.year}'
    path = f'{os.path.abspath(os.getcwd())}\excel'
    if not os.path.exists(path):
        os.makedirs(path)
    wb.save(f'{path}\staff_{today_date}.xlsx')


eel.init('web')
eel.start('index.html', size=(985, 830))
